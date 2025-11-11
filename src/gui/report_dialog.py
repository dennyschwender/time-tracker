"""
Dialog to select a date range and export an Excel report.
"""

from datetime import date
from pathlib import Path
from PyQt5.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QDateEdit,
    QFileDialog,
    QMessageBox,
)
from PyQt5.QtCore import Qt
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


class ReportDialog(QDialog):
    def __init__(self, entry_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Export Report")
        self.setModal(True)
        self.entry_manager = entry_manager

        layout = QVBoxLayout(self)

        # Date range inputs
        row = QHBoxLayout()
        row.addWidget(QLabel("Start date:"))
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(date.today())
        row.addWidget(self.start_date)

        row.addWidget(QLabel("End date:"))
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(date.today())
        row.addWidget(self.end_date)

        layout.addLayout(row)

        # Buttons
        btn_row = QHBoxLayout()
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.clicked.connect(self._on_export)
        btn_row.addWidget(self.export_btn)

        cancel_btn = QPushButton("Close")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)

        layout.addLayout(btn_row)

    def _on_export(self):
        start = self.start_date.date().toPyDate()
        end = self.end_date.date().toPyDate()

        # Ask for file path
        fn, _ = QFileDialog.getSaveFileName(self, "Save Report", str(Path.home() / "time_report.xlsx"), "Excel Files (*.xlsx)")
        if not fn:
            return

        try:
            dates, descriptions, matrix = self.entry_manager.generate_report(start, end)
            wb = Workbook()
            ws: Worksheet = wb.active  # type: ignore
            ws.title = "Time Report"

            # Header: first cell empty, then dates, then "Total"
            ws.cell(row=1, column=1, value="Description")
            for c, d in enumerate(dates, start=2):
                ws.cell(row=1, column=c, value=d.isoformat())
            # Add "Total" column header
            total_col = len(dates) + 2
            ws.cell(row=1, column=total_col, value="Total")

            # Fill rows with data and calculate row totals
            for r, desc in enumerate(descriptions, start=2):
                ws.cell(row=r, column=1, value=desc)
                row_vals = matrix.get(desc, [])
                row_total = 0.0
                for c, val in enumerate(row_vals, start=2):
                    ws.cell(row=r, column=c, value=val)
                    row_total += val
                # Write row total
                ws.cell(row=r, column=total_col, value=round(row_total, 2))

            # Add a row for column totals
            total_row = len(descriptions) + 2
            ws.cell(row=total_row, column=1, value="Total")
            
            # Calculate column totals (sum for each date)
            grand_total = 0.0
            for c, _ in enumerate(dates, start=2):
                col_total = 0.0
                for desc in descriptions:
                    row_vals = matrix.get(desc, [])
                    if c - 2 < len(row_vals):
                        col_total += row_vals[c - 2]
                ws.cell(row=total_row, column=c, value=round(col_total, 2))
                grand_total += col_total
            
            # Write grand total (bottom-right cell)
            ws.cell(row=total_row, column=total_col, value=round(grand_total, 2))

            # Auto-width for columns
            for idx, col in enumerate(ws.columns, start=1):
                max_length = 0
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(fn)
            QMessageBox.information(self, "Exported", f"Report saved to: {fn}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export report: {e}")
